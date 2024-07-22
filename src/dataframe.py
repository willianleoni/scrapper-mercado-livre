import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Alignment
from tqdm import tqdm

class DataFrameHandler:
    def __init__(self, products):
        self.df = self.products_to_dataframe(products)
    
    def products_to_dataframe(self, products):
        return pd.DataFrame(products)

    def add_discount_column(self):
        if 'price' in self.df.columns and 'discounted_price' in self.df.columns:
            self.df['price'] = self.df['price'].str.replace('.', '').astype(float, errors='ignore')
            self.df['discounted_price'] = self.df['discounted_price'].str.replace('.', '').astype(float, errors='ignore')
            self.df['discount'] = ((self.df['price'] - self.df['discounted_price']) / self.df['price']) * 100
        else:
            self.df['discount'] = None
        return self.df

    def convert_discount_to_percentage(self):
        self.df['discount'] = self.df['discount'].fillna(0)
        self.df['discount'] = self.df['discount'].apply(lambda x: str(int(x)) + '%' if not pd.isna(x) else '0%')
        return self.df

    def sort_by_discount(self):
        self.df['discount_numeric'] = self.df['discount'].str.rstrip('%').astype(float)
        self.df = self.df.sort_values(by='discount_numeric', ascending=False).drop(columns=['discount_numeric'])
        return self.df

    def save_to_excel(self):
        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        filename = f'../data/products_{timestamp}.xlsx'

        wb = Workbook()
        ws = wb.active
        ws.title = "Products"

        for r_idx, row in tqdm(enumerate(dataframe_to_rows(self.df, index=False, header=True), 1), desc="Writing to Excel"):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:  # Formata a linha de cabeçalho
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    if r_idx % 2 == 0:
                        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Cinza para linhas pares

        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        wb.save(filename)
        print(f"DataFrame saved and formatted in {filename}")